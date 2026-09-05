from fastapi import FastAPI, APIRouter, HTTPException, Header, Depends
from fastapi.responses import HTMLResponse, Response
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import logging
import uuid
import secrets
import calendar
import asyncio
import re
import bcrypt
import jwt
import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, BeforeValidator
from typing import List, Optional, Annotated, Any
from datetime import datetime, timezone, timedelta, date

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# ---------------------------------------------------------------------------
# Config / DB
# ---------------------------------------------------------------------------
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_SECRET = os.environ.get('JWT_SECRET', 'notifin-dev-secret')
JWT_ALG = 'HS256'
SESSION_DAYS = 7

# Emergent managed push
PUSH_BASE_URL = "https://integrations.emergentagent.com"
PUSH_KEY = os.environ.get("EMERGENT_PUSH_KEY", "placeholder")
EMERGENT_AUTH_URL = "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data"

FREE_PLAN_LIMIT = 3

# WhatsApp via Fonnte (simulation mode while token is empty)
FONNTE_TOKEN = os.environ.get("FONNTE_TOKEN", "")
FONNTE_BASE_URL = "https://api.fonnte.com"


def wa_live() -> bool:
    return bool(FONNTE_TOKEN.strip())


# Payment via Mayar.id (membership product "Notifin Premium").
# All four stay empty until KYC is approved and you fill them in on Railway —
# /auth/upgrade returns a clear "not configured yet" error until then, it
# never falls back to the old dummy toggle.
MAYAR_API_KEY = os.environ.get("MAYAR_API_KEY", "")
MAYAR_PRODUCT_ID = os.environ.get("MAYAR_PRODUCT_ID", "")
MAYAR_TIER_MONTHLY_ID = os.environ.get("MAYAR_TIER_MONTHLY_ID", "")
MAYAR_TIER_YEARLY_ID = os.environ.get("MAYAR_TIER_YEARLY_ID", "")
MAYAR_BASE_URL = "https://api.mayar.id"

# Mayar does not sign/HMAC its webhook body (confirmed against their public
# docs and a working third-party integration writeup — there is no header or
# payload field to check). The documented workaround, and what real Mayar
# integrations use, is a shared secret placed in the webhook URL itself:
# register "https://<backend>/api/webhooks/mayar?secret=<this value>" as the
# webhook URL in the Mayar dashboard, and this app rejects any call whose
# ?secret= doesn't match. Generate any long random string for it.
MAYAR_WEBHOOK_SECRET = os.environ.get("MAYAR_WEBHOOK_SECRET", "")


def mayar_live() -> bool:
    return bool(
        MAYAR_API_KEY.strip() and MAYAR_PRODUCT_ID.strip()
        and MAYAR_TIER_MONTHLY_ID.strip() and MAYAR_TIER_YEARLY_ID.strip()
    )

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

_push_client = httpx.AsyncClient(
    base_url=PUSH_BASE_URL,
    headers={"X-Push-Key": PUSH_KEY},
    timeout=10.0,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def new_user_id() -> str:
    return f"user_{uuid.uuid4().hex[:12]}"


def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode('utf-8'), hashed.encode('utf-8'))
    except Exception:
        return False


def make_session_token(user_id: str) -> str:
    payload = {
        "user_id": user_id,
        "exp": now_utc() + timedelta(days=SESSION_DAYS),
        "iat": now_utc(),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


async def persist_session(user_id: str, token: str):
    await db.user_sessions.insert_one({
        "session_token": token,
        "user_id": user_id,
        "created_at": now_utc().isoformat(),
        "expires_at": (now_utc() + timedelta(days=SESSION_DAYS)).isoformat(),
    })


def public_user(u: dict) -> dict:
    return {
        "user_id": u["user_id"],
        "email": u.get("email"),
        "name": u.get("name"),
        "picture": u.get("picture"),
        "plan": u.get("plan", "free"),
        "phone": u.get("phone"),
        "wa_live": wa_live(),
        "notify_channels": u.get("notify_channels", {"push": True, "whatsapp": False}),
        "monthly_limit": u.get("monthly_limit"),
    }


def normalize_phone(value: str) -> str:
    digits = re.sub(r"\D", "", value or "")
    if digits.startswith("0"):
        digits = "62" + digits[1:]
    if not re.fullmatch(r"62[1-9][0-9]{7,12}", digits):
        raise ValueError("invalid phone")
    return digits


def fmt_rp(v: float) -> str:
    return "Rp" + f"{round(v or 0):,}".replace(",", ".")


# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------
class RegisterBody(BaseModel):
    email: EmailStr
    password: str
    name: str


class LoginBody(BaseModel):
    email: EmailStr
    password: str


class SessionBody(BaseModel):
    session_id: str


class SubscriptionBody(BaseModel):
    name: str
    category: str = "other"
    price: float = 0
    billing_cycle: str = "monthly"          # monthly | yearly | weekly
    next_due_date: str                      # YYYY-MM-DD
    status: str = "paid"                    # trial | paid
    color: Optional[str] = None
    reminders: List[int] = Field(default_factory=lambda: [3, 1, 0])
    notes: Optional[str] = None
    registered_with: Optional[str] = None   # email/akun/no. HP dipakai daftar (opsional)


class RegisterPushBody(BaseModel):
    user_id: str
    platform: str
    device_token: str


# ---------------------------------------------------------------------------
# Auth dependency
# ---------------------------------------------------------------------------
async def get_current_user(authorization: Optional[str] = Header(None)) -> dict:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Missing token")
    token = authorization.split(" ", 1)[1].strip()
    session = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    exp = session.get("expires_at")
    if isinstance(exp, str):
        exp_dt = datetime.fromisoformat(exp)
    else:
        exp_dt = exp
    if exp_dt.tzinfo is None:
        exp_dt = exp_dt.replace(tzinfo=timezone.utc)
    if exp_dt < now_utc():
        raise HTTPException(status_code=401, detail="Session expired")
    user = await db.users.find_one({"user_id": session["user_id"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")

    # Throttled "last active" tracking for the admin panel — only write if
    # stale (>10 min) so this doesn't add a DB write to every single request.
    last_active = user.get("last_active_at")
    stale = True
    if last_active:
        try:
            stale = (now_utc() - datetime.fromisoformat(last_active)) > timedelta(minutes=10)
        except Exception:
            stale = True
    if stale:
        now_iso = now_utc().isoformat()
        await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"last_active_at": now_iso}})
        user["last_active_at"] = now_iso

    return user


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------
@api_router.post("/auth/register")
async def register(body: RegisterBody):
    existing = await db.users.find_one({"email": body.email.lower()})
    if existing:
        raise HTTPException(status_code=409, detail="Email sudah terdaftar")
    user = {
        "user_id": new_user_id(),
        "email": body.email.lower(),
        "name": body.name,
        "password_hash": hash_password(body.password),
        "picture": None,
        "plan": "free",
        "notify_channels": {"push": True, "whatsapp": False},
        "created_at": now_utc().isoformat(),
    }
    await db.users.insert_one(user)
    token = make_session_token(user["user_id"])
    await persist_session(user["user_id"], token)
    return {"session_token": token, "user": public_user(user)}


@api_router.post("/auth/login")
async def login(body: LoginBody):
    user = await db.users.find_one({"email": body.email.lower()}, {"_id": 0})
    if not user or not user.get("password_hash") or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Email atau password salah")
    token = make_session_token(user["user_id"])
    await persist_session(user["user_id"], token)
    return {"session_token": token, "user": public_user(user)}


@api_router.post("/auth/session")
async def google_session(body: SessionBody):
    try:
        resp = await httpx.AsyncClient(timeout=10.0).get(
            EMERGENT_AUTH_URL, headers={"X-Session-ID": body.session_id}
        )
    except Exception as e:
        logger.warning(f"Emergent auth error: {e}")
        raise HTTPException(status_code=401, detail="Auth gagal")
    if resp.status_code != 200:
        raise HTTPException(status_code=401, detail="Sesi tidak valid")
    data = resp.json()
    email = (data.get("email") or "").lower()
    existing = await db.users.find_one({"email": email}, {"_id": 0})
    if existing:
        user = existing
    else:
        user = {
            "user_id": new_user_id(),
            "email": email,
            "name": data.get("name"),
            "picture": data.get("picture"),
            "password_hash": None,
            "plan": "free",
            "notify_channels": {"push": True, "whatsapp": False},
            "created_at": now_utc().isoformat(),
        }
        await db.users.insert_one(user)
    token = data.get("session_token") or make_session_token(user["user_id"])
    await persist_session(user["user_id"], token)
    return {"session_token": token, "user": public_user(user)}


@api_router.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return {"user": public_user(user)}


@api_router.post("/auth/logout")
async def logout(authorization: Optional[str] = Header(None)):
    if authorization and authorization.lower().startswith("bearer "):
        token = authorization.split(" ", 1)[1].strip()
        await db.user_sessions.delete_one({"session_token": token})
    return {"status": "ok"}


class UpgradeBody(BaseModel):
    tier: str  # "monthly" | "yearly"


@api_router.post("/auth/upgrade")
async def upgrade(body: UpgradeBody, user: dict = Depends(get_current_user)):
    """Starts a real Mayar checkout — does NOT flip `plan` itself. `plan`
    only becomes "premium" once /webhooks/mayar confirms the payment."""
    if body.tier not in ("monthly", "yearly"):
        raise HTTPException(status_code=422, detail='tier harus "monthly" atau "yearly"')
    if not mayar_live():
        raise HTTPException(
            status_code=503,
            detail="Pembayaran belum aktif — masih menunggu verifikasi KYC Mayar selesai.",
        )
    tier_id = MAYAR_TIER_MONTHLY_ID if body.tier == "monthly" else MAYAR_TIER_YEARLY_ID
    checkout_link = await mayar_create_checkout(user, tier_id)
    if not checkout_link:
        raise HTTPException(
            status_code=502,
            detail="Mayar tidak mengembalikan link checkout. Cek log server untuk detail responsnya.",
        )
    return {"checkout_url": checkout_link}


async def mayar_create_checkout(user: dict, tier_id: str) -> Optional[str]:
    """Registers the user against a membership tier on Mayar, which — per
    Mayar's docs — creates a pending member + associated payment link for
    them to complete checkout. We only ever trust the webhook to actually
    grant premium; this call's response is used purely to get a URL to send
    the user to pay at.

    NOTE: Mayar's public docs for non-credit membership products are thin
    and, in places, inconsistent about the exact response shape here — this
    checks every field name we found evidence for. If Mayar's real response
    doesn't match any of them, this returns None and the raw response is
    logged so it can be fixed from a real response body once KYC is done.
    """
    try:
        async with httpx.AsyncClient(timeout=httpx.Timeout(20.0, connect=5.0)) as c:
            resp = await c.post(
                f"{MAYAR_BASE_URL}/hl/v2/memberships/members/create",
                headers={
                    "Authorization": f"Bearer {MAYAR_API_KEY.strip()}",
                    "Content-Type": "application/json",
                },
                json={
                    "productId": MAYAR_PRODUCT_ID.strip(),
                    "membershipTierId": tier_id.strip(),
                    "customerInfo": {
                        "name": user.get("name") or user["email"],
                        "email": user["email"],
                        "mobile": user.get("phone") or "-",
                    },
                    "membershipMonthlyPeriod": 1,
                },
            )
        result = resp.json()
    except Exception as e:
        logger.warning(f"Mayar checkout request failed: {e}")
        return None

    if resp.status_code >= 400:
        logger.warning(f"Mayar checkout rejected ({resp.status_code}): {result}")
        return None

    data = result.get("data", {}) or {}
    member = data.get("membershipCustomer", data)
    checkout_link = data.get("checkoutLink") or data.get("paymentLink") or member.get("checkoutLink")
    if not checkout_link:
        payment_link_id = member.get("paymentLinkId") or data.get("paymentLinkId")
        if payment_link_id:
            checkout_link = f"https://mayar.id/pl/checkout?product={payment_link_id}"
    if not checkout_link:
        logger.warning(f"Mayar checkout: no usable link in response: {result}")
    return checkout_link


# ---------------------------------------------------------------------------
# Mayar webhook — grants/revokes premium. `plan` is ONLY ever changed here
# (and by the self-service /auth/downgrade above), never by /auth/upgrade.
# ---------------------------------------------------------------------------
MAYAR_UPGRADE_EVENTS = {"membership.newMemberRegistered", "membership.changeTierMemberRegistered"}
MAYAR_DOWNGRADE_EVENTS = {"membership.memberExpired", "membership.memberUnsubscribed"}


async def find_user_by_mayar_customer(email: Optional[str], mobile: Optional[str]) -> Optional[dict]:
    email_norm = (email or "").strip().lower()
    if email_norm:
        u = await db.users.find_one({"email": email_norm}, {"_id": 0})
        if u:
            return u
    if mobile:
        try:
            phone_norm = normalize_phone(mobile)
        except ValueError:
            phone_norm = None
        if phone_norm:
            u = await db.users.find_one({"phone": phone_norm}, {"_id": 0})
            if u:
                return u
    return None


DEFAULT_PREMIUM_DAYS = 30  # fallback when Mayar's payload has no real expiry date


def parse_mayar_timestamp(value) -> Optional[str]:
    """Mayar's own doc examples show conflicting formats for date fields —
    epoch-milliseconds numbers in some, ISO strings in others. Handle both,
    return None (never raise) if it doesn't parse."""
    if value is None:
        return None
    try:
        if isinstance(value, (int, float)):
            return datetime.fromtimestamp(value / 1000, tz=timezone.utc).isoformat()
        if isinstance(value, str) and value.strip():
            return datetime.fromisoformat(value.replace("Z", "+00:00")).isoformat()
    except Exception:
        return None
    return None


async def process_mayar_event(event: str, data: dict) -> dict:
    """Shared by the real webhook and /test/simulate-mayar-webhook, so both
    exercise the identical matching/update/logging logic — never trust a
    payload blindly: this only acts on event names we recognize, and only
    updates a user it can actually match by email or phone."""
    membership_customer = data.get("membershipCustomer") or {}
    customer_email = data.get("customerEmail") or membership_customer.get("customerEmail")
    customer_mobile = data.get("customerMobile") or membership_customer.get("customerMobile")

    matched_user = await find_user_by_mayar_customer(customer_email, customer_mobile)
    is_upgrade = event in MAYAR_UPGRADE_EVENTS or (
        event == "payment.received" and bool(membership_customer)
    )
    is_downgrade = event in MAYAR_DOWNGRADE_EVENTS

    if matched_user and is_upgrade:
        action = "upgraded_to_premium"
        expires_at = (
            parse_mayar_timestamp(membership_customer.get("expiredAt"))
            or parse_mayar_timestamp(membership_customer.get("nextPayment"))
            or (now_utc() + timedelta(days=DEFAULT_PREMIUM_DAYS)).isoformat()
        )
        await db.users.update_one(
            {"user_id": matched_user["user_id"]},
            {"$set": {
                "plan": "premium",
                "premium_since": now_utc().isoformat(),
                "premium_expires_at": expires_at,
            }},
        )
    elif matched_user and is_downgrade:
        action = "downgraded_to_free"
        await db.users.update_one(
            {"user_id": matched_user["user_id"]},
            {"$set": {"plan": "free", "premium_expires_at": None}},
        )
    elif not matched_user and (is_upgrade or is_downgrade):
        action = "no_matching_user"
    else:
        action = "ignored_event"

    log_entry = {
        "id": str(uuid.uuid4()),
        "event": event,
        "customer_email": customer_email,
        "customer_mobile": customer_mobile,
        "matched_user_id": matched_user["user_id"] if matched_user else None,
        "action": action,
        "received_at": now_utc().isoformat(),
    }
    await db.mayar_webhook_log.insert_one(log_entry)
    logger.info(
        f"Mayar webhook: event={event} action={action} user={log_entry['matched_user_id']}"
    )
    return {"action": action, "matched_user_id": log_entry["matched_user_id"]}


@api_router.post("/webhooks/mayar")
async def mayar_webhook(payload: Optional[dict] = None, secret: Optional[str] = None):
    # Mayar has no signature/HMAC scheme (see MAYAR_WEBHOOK_SECRET comment
    # above) — this shared-secret query param is the only line of defense,
    # so it's checked before anything else, and both a missing configured
    # secret and a mismatched one are rejected identically.
    if not MAYAR_WEBHOOK_SECRET.strip() or secret != MAYAR_WEBHOOK_SECRET.strip():
        raise HTTPException(status_code=401, detail="Invalid or missing webhook secret")
    body = payload or {}
    event = body.get("event", "")
    data = body.get("data") or {}
    result = await process_mayar_event(event, data)
    return {"status": "ok", **result}


class SimulateMayarWebhookBody(BaseModel):
    event: str = "membership.newMemberRegistered"


@api_router.post("/test/simulate-mayar-webhook")
async def test_simulate_mayar_webhook(
    body: SimulateMayarWebhookBody, user: dict = Depends(get_current_user)
):
    """TESTING ONLY — runs the exact same code path as the real webhook
    (process_mayar_event), but always targets the CALLER's own account
    (customerEmail is forced to your logged-in email, never something you
    pass in), so this can't be used to flip anyone else's plan. Use it to
    verify event routing + user matching + plan flip + the debug log before
    Mayar's real webhook exists. Valid `event` values: membership.newMemberRegistered,
    membership.changeTierMemberRegistered, membership.memberExpired,
    membership.memberUnsubscribed.
    """
    data = {
        "customerEmail": user["email"],
        "customerMobile": user.get("phone"),
        "customerName": user.get("name"),
    }
    result = await process_mayar_event(body.event, data)
    return {"status": "ok", "simulated_event": body.event, **result}


@api_router.post("/auth/downgrade")
async def mock_downgrade(user: dict = Depends(get_current_user)):
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"plan": "free"}})
    updated = await db.users.find_one({"user_id": user["user_id"]}, {"_id": 0})
    return {"user": public_user(updated)}


class ChannelsBody(BaseModel):
    push: bool = True
    whatsapp: bool = False


@api_router.put("/auth/channels")
async def update_channels(body: ChannelsBody, user: dict = Depends(get_current_user)):
    await db.users.update_one({"user_id": user["user_id"]},
                              {"$set": {"notify_channels": body.model_dump()}})
    updated = await db.users.find_one({"user_id": user["user_id"]}, {"_id": 0})
    return {"user": public_user(updated)}


class PhoneBody(BaseModel):
    phone: str


@api_router.put("/auth/phone")
async def update_phone(body: PhoneBody, user: dict = Depends(get_current_user)):
    raw = body.phone.strip()
    if raw == "":
        normalized = None
    else:
        try:
            normalized = normalize_phone(raw)
        except ValueError:
            raise HTTPException(status_code=422,
                                detail="Nomor WhatsApp tidak valid. Pakai format 08xx atau +62xx")
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"phone": normalized}})
    updated = await db.users.find_one({"user_id": user["user_id"]}, {"_id": 0})
    return {"user": public_user(updated)}


class LimitBody(BaseModel):
    monthly_limit: Optional[float] = None   # null/0 = tanpa limit


@api_router.put("/auth/limit")
async def update_limit(body: LimitBody, user: dict = Depends(get_current_user)):
    value = body.monthly_limit
    if value is not None and value <= 0:
        raise HTTPException(status_code=422, detail="Limit harus lebih dari 0")
    await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"monthly_limit": value}})
    updated = await db.users.find_one({"user_id": user["user_id"]}, {"_id": 0})
    return {"user": public_user(updated)}


# ---------------------------------------------------------------------------
# Subscriptions
# ---------------------------------------------------------------------------
def sub_public(s: dict) -> dict:
    return {
        "id": s["id"],
        "name": s["name"],
        "category": s.get("category", "other"),
        "price": s.get("price", 0),
        "billing_cycle": s.get("billing_cycle", "monthly"),
        "next_due_date": s.get("next_due_date"),
        "status": s.get("status", "paid"),
        "color": s.get("color"),
        "reminders": s.get("reminders", [3, 1, 0]),
        "notes": s.get("notes"),
        "registered_with": s.get("registered_with"),
        "created_at": s.get("created_at"),
    }


def monthly_cost(s: dict) -> float:
    price = float(s.get("price", 0) or 0)
    cycle = s.get("billing_cycle", "monthly")
    if cycle == "yearly":
        return price / 12.0
    if cycle == "weekly":
        return price * 52.0 / 12.0
    return price


async def active_count(user_id: str) -> int:
    return await db.subscriptions.count_documents(
        {"user_id": user_id, "deleted_at": None}
    )


async def ensure_distinct_registered_with(
    user_id: str, name: str, registered_with: Optional[str], exclude_id: Optional[str] = None,
):
    """Kalau ada langganan lain dengan nama sama (case-insensitive), wajib isi
    `registered_with` yang berbeda supaya keduanya bisa dibedakan."""
    q: dict = {"user_id": user_id, "deleted_at": None}
    if exclude_id:
        q["id"] = {"$ne": exclude_id}
    existing = await db.subscriptions.find(q, {"_id": 0}).to_list(500)
    name_norm = name.strip().lower()
    dupes = [d for d in existing if (d.get("name") or "").strip().lower() == name_norm]
    if not dupes:
        return
    rw = (registered_with or "").strip()
    if not rw:
        raise HTTPException(
            status_code=422,
            detail=f'Sudah ada langganan "{name}" lain. Isi "Terdaftar dengan" biar bisa dibedakan.',
        )
    rw_norm = rw.lower()
    for d in dupes:
        other_rw = (d.get("registered_with") or "").strip().lower()
        if other_rw and other_rw == rw_norm:
            raise HTTPException(
                status_code=422,
                detail=f'Akun "{rw}" sudah dipakai untuk langganan "{name}" lainnya. Pakai akun yang berbeda.',
            )


@api_router.get("/subscriptions")
async def list_subscriptions(
    category: Optional[str] = None,
    status: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    q: dict = {"user_id": user["user_id"], "deleted_at": None}
    if category and category != "all":
        q["category"] = category
    if status and status != "all":
        q["status"] = status
    docs = await db.subscriptions.find(q, {"_id": 0}).sort("next_due_date", 1).to_list(500)
    return {"subscriptions": [sub_public(d) for d in docs]}


@api_router.post("/subscriptions")
async def create_subscription(body: SubscriptionBody, user: dict = Depends(get_current_user)):
    if user.get("plan", "free") == "free":
        count = await active_count(user["user_id"])
        if count >= FREE_PLAN_LIMIT:
            raise HTTPException(
                status_code=403,
                detail={"code": "limit_reached",
                        "message": f"Paket gratis maksimal {FREE_PLAN_LIMIT} langganan aktif."},
            )
    await ensure_distinct_registered_with(user["user_id"], body.name, body.registered_with)
    doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["user_id"],
        **body.model_dump(),
        "deleted_at": None,
        "created_at": now_utc().isoformat(),
        "updated_at": now_utc().isoformat(),
    }
    await db.subscriptions.insert_one(doc)
    return {"subscription": sub_public(doc)}


@api_router.get("/subscriptions/{sub_id}")
async def get_subscription(sub_id: str, user: dict = Depends(get_current_user)):
    doc = await db.subscriptions.find_one(
        {"id": sub_id, "user_id": user["user_id"], "deleted_at": None}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Langganan tidak ditemukan")
    return {"subscription": sub_public(doc)}


@api_router.put("/subscriptions/{sub_id}")
async def update_subscription(sub_id: str, body: SubscriptionBody, user: dict = Depends(get_current_user)):
    doc = await db.subscriptions.find_one(
        {"id": sub_id, "user_id": user["user_id"], "deleted_at": None})
    if not doc:
        raise HTTPException(status_code=404, detail="Langganan tidak ditemukan")
    await ensure_distinct_registered_with(
        user["user_id"], body.name, body.registered_with, exclude_id=sub_id)
    update = {**body.model_dump(), "updated_at": now_utc().isoformat()}
    await db.subscriptions.update_one({"id": sub_id}, {"$set": update})
    updated = await db.subscriptions.find_one({"id": sub_id}, {"_id": 0})
    return {"subscription": sub_public(updated)}


@api_router.delete("/subscriptions/{sub_id}")
async def delete_subscription(sub_id: str, user: dict = Depends(get_current_user)):
    doc = await db.subscriptions.find_one(
        {"id": sub_id, "user_id": user["user_id"], "deleted_at": None})
    if not doc:
        raise HTTPException(status_code=404, detail="Langganan tidak ditemukan")
    await db.subscriptions.update_one(
        {"id": sub_id}, {"$set": {"deleted_at": now_utc().isoformat()}})
    return {"status": "deleted"}


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------
@api_router.get("/dashboard")
async def dashboard(user: dict = Depends(get_current_user)):
    docs = await db.subscriptions.find(
        {"user_id": user["user_id"], "deleted_at": None}, {"_id": 0}).to_list(500)

    total_monthly = 0.0
    by_cat: dict = {}
    upcoming = []
    ending_trials = []
    most_exp = None
    most_exp_cost = 0.0
    today = date.today()
    horizon = today + timedelta(days=7)

    for d in docs:
        m = monthly_cost(d)
        total_monthly += m
        if m > most_exp_cost:
            most_exp_cost = m
            most_exp = d
        cat = d.get("category", "other")
        by_cat.setdefault(cat, {"category": cat, "total": 0.0, "count": 0})
        by_cat[cat]["total"] += m
        by_cat[cat]["count"] += 1

        due_raw = d.get("next_due_date")
        try:
            due = date.fromisoformat(due_raw)
        except Exception:
            due = None
        if due is not None:
            days_left = (due - today).days
            if today <= due <= horizon:
                pub = sub_public(d)
                pub["days_left"] = days_left
                upcoming.append(pub)
            if d.get("status") == "trial" and 0 <= days_left <= 14:
                pub = sub_public(d)
                pub["days_left"] = days_left
                ending_trials.append(pub)

    upcoming.sort(key=lambda x: x.get("days_left", 99))
    ending_trials.sort(key=lambda x: x.get("days_left", 99))
    by_category = sorted(by_cat.values(), key=lambda x: x["total"], reverse=True)

    # Real spending history (for the spending chart): record/refresh a snapshot
    # of *this* month's total every time the dashboard is viewed. Past months'
    # snapshots are never touched again once the month has moved on, so they
    # become the permanent historical record — data starts accumulating from
    # today, not reconstructed/estimated from subscription start dates.
    await db.spending_snapshots.update_one(
        {"user_id": user["user_id"], "period": today.strftime("%Y-%m")},
        {"$set": {"total": round(total_monthly), "updated_at": now_utc().isoformat()}},
        upsert=True,
    )

    return {
        "total_this_month": round(total_monthly),
        "projection_next_month": round(total_monthly),
        "active_count": len(docs),
        "plan": user.get("plan", "free"),
        "free_limit": FREE_PLAN_LIMIT,
        "upcoming": upcoming,
        "most_expensive": (
            {**sub_public(most_exp), "monthly_cost": round(most_exp_cost)}
            if most_exp is not None and most_exp_cost > 0 else None
        ),
        "ending_trials": ending_trials,
        "by_category": [
            {"category": c["category"], "total": round(c["total"]), "count": c["count"]}
            for c in by_category
        ],
    }


@api_router.get("/analytics/spending")
async def spending_history(range: str = "monthly", user: dict = Depends(get_current_user)):
    """Real, tracked spending history — each row is a snapshot taken the last
    time the dashboard was loaded during that month (see /dashboard above).
    Nothing is backfilled/estimated: history only exists from the point this
    feature shipped onward, and grows one entry per month as time passes."""
    snapshots = await db.spending_snapshots.find(
        {"user_id": user["user_id"]}, {"_id": 0}).sort("period", 1).to_list(240)

    if range == "yearly":
        by_year: dict = {}
        for s in snapshots:
            year = s["period"][:4]
            by_year.setdefault(year, 0)
            by_year[year] += s.get("total", 0)
        points = [{"period": y, "total": round(t)} for y, t in sorted(by_year.items())]
    else:
        points = [{"period": s["period"], "total": round(s.get("total", 0))} for s in snapshots]

    return {
        "range": "yearly" if range == "yearly" else "monthly",
        "points": points,
        "tracking_since": snapshots[0]["period"] if snapshots else None,
    }


# ---------------------------------------------------------------------------
# Fase 2: Groups (Family/Team Sharing)
# ---------------------------------------------------------------------------
CODE_ALPHABET = "ABCDEFGHJKMNPQRSTUVWXYZ23456789"


def gen_invite_code() -> str:
    return "".join(secrets.choice(CODE_ALPHABET) for _ in range(6))


def add_cycle(d: date, cycle: str) -> date:
    if cycle == "weekly":
        return d + timedelta(days=7)
    if cycle == "yearly":
        try:
            return d.replace(year=d.year + 1)
        except ValueError:
            return d.replace(year=d.year + 1, day=28)
    y, m = d.year, d.month + 1
    if m > 12:
        y, m = y + 1, 1
    day = min(d.day, calendar.monthrange(y, m)[1])
    return date(y, m, day)


async def advance_group_sub(s: dict) -> dict:
    """Auto-advance due date past today; payments are keyed per due date so
    statuses reset automatically each new billing period."""
    try:
        due = date.fromisoformat(s.get("next_due_date"))
    except Exception:
        return s
    today = date.today()
    changed = False
    while due < today:
        due = add_cycle(due, s.get("billing_cycle", "monthly"))
        changed = True
    if changed:
        s["next_due_date"] = due.isoformat()
        await db.group_subscriptions.update_one(
            {"id": s["id"]}, {"$set": {"next_due_date": s["next_due_date"]}})
    return s


def compute_splits(s: dict, members: List[dict], period: Optional[str] = None) -> List[dict]:
    period = period or s.get("next_due_date") or ""
    payments = (s.get("payments") or {}).get(period, {})
    splits = []
    if s.get("split_type") == "custom":
        cs = s.get("custom_splits") or {}
        for m in members:
            splits.append({
                "user_id": m["user_id"],
                "name": m.get("name") or "Anggota",
                "amount": round(float(cs.get(m["user_id"], 0) or 0)),
                "paid": bool(payments.get(m["user_id"])),
            })
    else:
        n = max(len(members), 1)
        share = float(s.get("price", 0) or 0) / n
        for m in members:
            splits.append({
                "user_id": m["user_id"],
                "name": m.get("name") or "Anggota",
                "amount": round(share),
                "paid": bool(payments.get(m["user_id"])),
            })
    return splits


def group_sub_public(s: dict) -> dict:
    return {
        "id": s["id"],
        "name": s["name"],
        "category": s.get("category", "other"),
        "price": s.get("price", 0),
        "billing_cycle": s.get("billing_cycle", "monthly"),
        "next_due_date": s.get("next_due_date"),
        "split_type": s.get("split_type", "equal"),
        "custom_splits": s.get("custom_splits"),
        "created_at": s.get("created_at"),
    }


class GroupBody(BaseModel):
    name: str


class JoinBody(BaseModel):
    code: str


class GroupSubBody(BaseModel):
    name: str
    category: str = "other"
    price: float = 0
    billing_cycle: str = "monthly"
    next_due_date: str
    split_type: str = "equal"               # equal | custom
    custom_splits: Optional[dict] = None    # {user_id: amount}


class PayBody(BaseModel):
    user_id: Optional[str] = None
    paid: bool = True


async def get_group_for_member(gid: str, user_id: str) -> dict:
    g = await db.groups.find_one({"id": gid}, {"_id": 0})
    if not g:
        raise HTTPException(status_code=404, detail="Grup tidak ditemukan")
    if not any(m["user_id"] == user_id for m in g.get("members", [])):
        raise HTTPException(status_code=403, detail="Kamu bukan anggota grup ini")
    return g


@api_router.post("/groups")
async def create_group(body: GroupBody, user: dict = Depends(get_current_user)):
    if user.get("plan", "free") != "premium":
        raise HTTPException(
            status_code=403,
            detail={"code": "premium_required",
                    "message": "Buat grup adalah fitur Premium. Semua orang tetap bisa gabung lewat kode."},
        )
    if not body.name.strip():
        raise HTTPException(status_code=422, detail="Nama grup wajib diisi")
    code = gen_invite_code()
    for _ in range(5):
        if not await db.groups.find_one({"invite_code": code}):
            break
        code = gen_invite_code()
    g = {
        "id": str(uuid.uuid4()),
        "name": body.name.strip(),
        "owner_id": user["user_id"],
        "invite_code": code,
        "members": [{"user_id": user["user_id"], "name": user.get("name"),
                     "joined_at": now_utc().isoformat()}],
        "created_at": now_utc().isoformat(),
    }
    await db.groups.insert_one(g)
    g.pop("_id", None)
    return {"group": g}


@api_router.get("/groups")
async def list_groups(user: dict = Depends(get_current_user)):
    docs = await db.groups.find(
        {"members.user_id": user["user_id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)
    out = []
    for g in docs:
        subs = await db.group_subscriptions.find(
            {"group_id": g["id"], "deleted_at": None}, {"_id": 0}).to_list(200)
        my_share = 0.0
        total_price = 0.0
        for s in subs:
            total_price += float(s.get("price", 0) or 0)
            for sp in compute_splits(s, g.get("members", [])):
                if sp["user_id"] == user["user_id"]:
                    my_share += sp["amount"]
        out.append({
            "id": g["id"],
            "name": g["name"],
            "invite_code": g["invite_code"],
            "is_owner": g["owner_id"] == user["user_id"],
            "member_count": len(g.get("members", [])),
            "sub_count": len(subs),
            "my_share": round(my_share),
            "total_price": round(total_price),
        })
    return {"groups": out}


@api_router.post("/groups/join")
async def join_group(body: JoinBody, user: dict = Depends(get_current_user)):
    code = body.code.strip().upper()
    g = await db.groups.find_one({"invite_code": code}, {"_id": 0})
    if not g:
        raise HTTPException(status_code=404, detail="Kode grup tidak ditemukan")
    if any(m["user_id"] == user["user_id"] for m in g.get("members", [])):
        raise HTTPException(status_code=409, detail="Kamu sudah jadi anggota grup ini")
    await db.groups.update_one(
        {"id": g["id"]},
        {"$push": {"members": {"user_id": user["user_id"], "name": user.get("name"),
                               "joined_at": now_utc().isoformat()}}})
    return {"status": "joined", "group_id": g["id"], "name": g["name"]}


@api_router.get("/groups/{gid}")
async def group_detail(gid: str, user: dict = Depends(get_current_user)):
    g = await get_group_for_member(gid, user["user_id"])
    members = g.get("members", [])
    subs_docs = await db.group_subscriptions.find(
        {"group_id": gid, "deleted_at": None}, {"_id": 0}).sort("next_due_date", 1).to_list(200)

    subs = []
    unpaid_names: set = set()
    total_price = 0.0
    my_total = 0.0
    for s in subs_docs:
        s = await advance_group_sub(s)
        splits = compute_splits(s, members)
        unpaid = [sp for sp in splits if not sp["paid"] and sp["amount"] > 0]
        for sp in unpaid:
            unpaid_names.add(sp["name"])
        mine = next((sp for sp in splits if sp["user_id"] == user["user_id"]), None)
        total_price += float(s.get("price", 0) or 0)
        if mine:
            my_total += mine["amount"]
        subs.append({
            **group_sub_public(s),
            "splits": splits,
            "unpaid_count": len(unpaid),
            "my_amount": mine["amount"] if mine else 0,
            "my_paid": mine["paid"] if mine else False,
        })

    return {"group": {
        "id": g["id"],
        "name": g["name"],
        "owner_id": g["owner_id"],
        "invite_code": g["invite_code"],
        "is_owner": g["owner_id"] == user["user_id"],
        "members": [{**m, "is_owner": m["user_id"] == g["owner_id"]} for m in members],
        "subscriptions": subs,
        "unpaid_members": sorted(unpaid_names),
        "total_price": round(total_price),
        "my_total": round(my_total),
        "created_at": g.get("created_at"),
    }}


@api_router.post("/groups/{gid}/leave")
async def leave_group(gid: str, user: dict = Depends(get_current_user)):
    g = await get_group_for_member(gid, user["user_id"])
    if g["owner_id"] == user["user_id"]:
        raise HTTPException(status_code=400,
                            detail="Koordinator tidak bisa keluar. Hapus grup jika sudah tidak dipakai.")
    await db.groups.update_one(
        {"id": gid}, {"$pull": {"members": {"user_id": user["user_id"]}}})
    return {"status": "left"}


@api_router.delete("/groups/{gid}")
async def delete_group(gid: str, user: dict = Depends(get_current_user)):
    g = await get_group_for_member(gid, user["user_id"])
    if g["owner_id"] != user["user_id"]:
        raise HTTPException(status_code=403, detail="Hanya koordinator yang bisa menghapus grup")
    await db.group_subscriptions.delete_many({"group_id": gid})
    await db.groups.delete_one({"id": gid})
    return {"status": "deleted"}


@api_router.post("/groups/{gid}/subscriptions")
async def create_group_sub(gid: str, body: GroupSubBody, user: dict = Depends(get_current_user)):
    g = await get_group_for_member(gid, user["user_id"])
    if g["owner_id"] != user["user_id"]:
        raise HTTPException(status_code=403, detail="Hanya koordinator yang bisa menambah langganan grup")
    doc = {
        "id": str(uuid.uuid4()),
        "group_id": gid,
        **body.model_dump(),
        "payments": {},
        "deleted_at": None,
        "created_at": now_utc().isoformat(),
        "updated_at": now_utc().isoformat(),
    }
    await db.group_subscriptions.insert_one(doc)
    return {"subscription": group_sub_public(doc)}


@api_router.put("/groups/{gid}/subscriptions/{sid}")
async def update_group_sub(gid: str, sid: str, body: GroupSubBody,
                           user: dict = Depends(get_current_user)):
    g = await get_group_for_member(gid, user["user_id"])
    if g["owner_id"] != user["user_id"]:
        raise HTTPException(status_code=403, detail="Hanya koordinator yang bisa mengubah langganan grup")
    doc = await db.group_subscriptions.find_one(
        {"id": sid, "group_id": gid, "deleted_at": None})
    if not doc:
        raise HTTPException(status_code=404, detail="Langganan grup tidak ditemukan")
    await db.group_subscriptions.update_one(
        {"id": sid}, {"$set": {**body.model_dump(), "updated_at": now_utc().isoformat()}})
    updated = await db.group_subscriptions.find_one({"id": sid}, {"_id": 0})
    return {"subscription": group_sub_public(updated)}


@api_router.delete("/groups/{gid}/subscriptions/{sid}")
async def delete_group_sub(gid: str, sid: str, user: dict = Depends(get_current_user)):
    g = await get_group_for_member(gid, user["user_id"])
    if g["owner_id"] != user["user_id"]:
        raise HTTPException(status_code=403, detail="Hanya koordinator yang bisa menghapus langganan grup")
    res = await db.group_subscriptions.update_one(
        {"id": sid, "group_id": gid, "deleted_at": None},
        {"$set": {"deleted_at": now_utc().isoformat()}})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Langganan grup tidak ditemukan")
    return {"status": "deleted"}


@api_router.put("/groups/{gid}/subscriptions/{sid}/pay")
async def pay_group_sub(gid: str, sid: str, body: PayBody,
                        user: dict = Depends(get_current_user)):
    g = await get_group_for_member(gid, user["user_id"])
    target = body.user_id or user["user_id"]
    if target != user["user_id"] and g["owner_id"] != user["user_id"]:
        raise HTTPException(status_code=403,
                            detail="Hanya koordinator yang bisa mengubah status bayar anggota lain")
    if not any(m["user_id"] == target for m in g.get("members", [])):
        raise HTTPException(status_code=404, detail="Anggota tidak ditemukan")
    s = await db.group_subscriptions.find_one(
        {"id": sid, "group_id": gid, "deleted_at": None}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Langganan grup tidak ditemukan")
    s = await advance_group_sub(s)
    period = s["next_due_date"]
    await db.group_subscriptions.update_one(
        {"id": sid}, {"$set": {f"payments.{period}.{target}": body.paid}})
    return {"status": "ok", "period": period, "user_id": target, "paid": body.paid}


# ---------------------------------------------------------------------------
# Fase 3: WhatsApp (Fonnte) + reminder scheduler + nudge + payment history
# ---------------------------------------------------------------------------
async def send_whatsapp(phone: str, message: str) -> dict:
    record = {"phone": phone, "message": message, "created_at": now_utc().isoformat()}
    if not wa_live():
        record.update({"simulated": True, "status": "simulated"})
        await db.wa_outbox.insert_one(record)
        logger.info(f"[WA SIMULASI] -> {phone}: {message}")
        return {"status": True, "simulated": True}
    try:
        async with httpx.AsyncClient(timeout=httpx.Timeout(15.0, connect=5.0)) as c:
            resp = await c.post(
                f"{FONNTE_BASE_URL}/send",
                headers={"Authorization": FONNTE_TOKEN.strip()},
                data={"target": phone, "message": message, "countryCode": "0"},
            )
        result = resp.json()
        ok = result.get("status") is True
        record.update({"simulated": False,
                       "status": "sent" if ok else "failed", "fonnte": result})
        await db.wa_outbox.insert_one(record)
        if not ok:
            logger.warning(f"Fonnte rejected: {result.get('reason')}")
        return {"status": ok, "simulated": False}
    except Exception as e:
        record.update({"simulated": False, "status": "error", "error": str(e)})
        await db.wa_outbox.insert_one(record)
        logger.warning(f"Fonnte send failed: {e}")
        return {"status": False, "simulated": False}


def when_label(offset: int) -> str:
    if offset == 0:
        return "hari ini"
    if offset == 1:
        return "besok"
    return f"{offset} hari lagi (H-{offset})"


async def claim_notif(key: str) -> bool:
    """Idempotency claim — returns False if this notification was already sent."""
    try:
        await db.notif_log.insert_one({"key": key, "created_at": now_utc().isoformat()})
        return True
    except Exception:
        return False


async def reminder_sweep():
    today = date.today()

    # Personal subscriptions -> WhatsApp for premium users with WA enabled + phone.
    users = await db.users.find(
        {"plan": "premium", "notify_channels.whatsapp": True,
         "phone": {"$nin": [None, ""]}}, {"_id": 0}).to_list(1000)
    for u in users:
        subs = await db.subscriptions.find(
            {"user_id": u["user_id"], "deleted_at": None}, {"_id": 0}).to_list(500)
        for s in subs:
            try:
                due = date.fromisoformat(s.get("next_due_date"))
            except Exception:
                continue
            offset = (due - today).days
            if offset not in (s.get("reminders") or []):
                continue
            key = f"wa:personal:{s['id']}:{s['next_due_date']}:{offset}"
            if await claim_notif(key):
                msg = (f"Halo {u.get('name') or 'kamu'}! 🔔 Langganan {s['name']} kamu "
                       f"{fmt_rp(s.get('price', 0))} jatuh tempo {when_label(offset)}. "
                       f"Jangan lupa bayar atau cancel ya — Notifin")
                await send_whatsapp(u["phone"], msg)

    # Group subscriptions -> push to unpaid members, WA to eligible unpaid members.
    gsubs = await db.group_subscriptions.find(
        {"deleted_at": None}, {"_id": 0}).to_list(2000)
    for s in gsubs:
        s = await advance_group_sub(s)
        try:
            due = date.fromisoformat(s.get("next_due_date"))
        except Exception:
            continue
        offset = (due - today).days
        if offset not in (3, 1, 0):
            continue
        g = await db.groups.find_one({"id": s["group_id"]}, {"_id": 0})
        if not g:
            continue
        for sp in compute_splits(s, g.get("members", [])):
            if sp["paid"] or sp["amount"] <= 0:
                continue
            uid = sp["user_id"]
            body_text = (f"Bagianmu {fmt_rp(sp['amount'])} untuk {s['name']} di grup "
                         f"\"{g['name']}\" jatuh tempo {when_label(offset)}.")
            if await claim_notif(f"push:group:{s['id']}:{s['next_due_date']}:{offset}:{uid}"):
                try:
                    await send_push([uid], {"title": "Tagihan grup 🔔", "message": body_text})
                except Exception as e:
                    logger.info(f"group push skipped: {e}")
            member = await db.users.find_one({"user_id": uid}, {"_id": 0})
            if (member and member.get("plan") == "premium"
                    and member.get("notify_channels", {}).get("whatsapp")
                    and member.get("phone")):
                if await claim_notif(f"wa:group:{s['id']}:{s['next_due_date']}:{offset}:{uid}"):
                    msg = (f"Halo {member.get('name')}! 🔔 {body_text} "
                           f"Jangan lupa bayar ya — Notifin")
                    await send_whatsapp(member["phone"], msg)


async def scheduler_loop():
    await asyncio.sleep(10)
    while True:
        try:
            await reminder_sweep()
        except Exception as e:
            logger.warning(f"reminder sweep failed: {e}")
        await asyncio.sleep(1800)


class NudgeBody(BaseModel):
    user_id: str


@api_router.post("/groups/{gid}/subscriptions/{sid}/nudge")
async def nudge_member(gid: str, sid: str, body: NudgeBody,
                       user: dict = Depends(get_current_user)):
    g = await get_group_for_member(gid, user["user_id"])
    if g["owner_id"] != user["user_id"]:
        raise HTTPException(status_code=403, detail="Hanya koordinator yang bisa mengingatkan anggota")
    if body.user_id == user["user_id"]:
        raise HTTPException(status_code=400, detail="Tidak bisa mengingatkan diri sendiri")
    if not any(m["user_id"] == body.user_id for m in g.get("members", [])):
        raise HTTPException(status_code=404, detail="Anggota tidak ditemukan")
    s = await db.group_subscriptions.find_one(
        {"id": sid, "group_id": gid, "deleted_at": None}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=404, detail="Langganan grup tidak ditemukan")
    s = await advance_group_sub(s)
    sp = next((x for x in compute_splits(s, g.get("members", []))
               if x["user_id"] == body.user_id), None)
    if not sp:
        raise HTTPException(status_code=404, detail="Anggota tidak ditemukan")
    if sp["paid"]:
        raise HTTPException(status_code=400, detail="Anggota ini sudah bayar")

    key = f"nudge:{sid}:{s['next_due_date']}:{body.user_id}:{date.today().isoformat()}"
    if not await claim_notif(key):
        raise HTTPException(status_code=429,
                            detail="Sudah diingatkan hari ini. Coba lagi besok ya.")

    channels = []
    body_text = (f"Bagianmu {fmt_rp(sp['amount'])} untuk {s['name']} di grup "
                 f"\"{g['name']}\" belum dibayar. Yuk segera lunasi!")
    try:
        await send_push([body.user_id],
                        {"title": f"{user.get('name')} mengingatkan 👋", "message": body_text})
        channels.append("push")
    except Exception as e:
        logger.info(f"nudge push skipped: {e}")
    target = await db.users.find_one({"user_id": body.user_id}, {"_id": 0})
    if target and target.get("phone"):
        res = await send_whatsapp(
            target["phone"],
            f"Halo {target.get('name')}! 👋 {user.get('name')} mengingatkan: {body_text} — Notifin")
        if res.get("status"):
            channels.append("whatsapp")
    return {"status": "sent", "channels": channels, "wa_simulated": not wa_live()}


# ---------------------------------------------------------------------------
# TESTING ONLY — manually fire a WhatsApp reminder for one subscription.
# Does not touch reminder_sweep()/scheduler_loop() or their day-offset
# (H-3/H-1/H-0) and notif_log dedup rules at all — this bypasses all of
# that on purpose so you can test sending without waiting for the
# scheduler or hitting the "already sent today" guard. Safe to leave in:
# while FONNTE_TOKEN is unset, send_whatsapp() stays in simulation mode
# (logs + wa_outbox only, no real message goes out).
# ---------------------------------------------------------------------------
class TestReminderBody(BaseModel):
    subscription_id: str


@api_router.post("/test/send-reminder")
async def test_send_reminder(body: TestReminderBody, user: dict = Depends(get_current_user)):
    sub = await db.subscriptions.find_one(
        {"id": body.subscription_id, "user_id": user["user_id"], "deleted_at": None}, {"_id": 0})
    if not sub:
        raise HTTPException(status_code=404, detail="Langganan tidak ditemukan")
    if not user.get("phone"):
        raise HTTPException(status_code=422, detail="Nomor WhatsApp belum diatur di akun ini")

    msg = (f"[TEST] Halo {user.get('name') or 'kamu'}! 🔔 Langganan {sub['name']} kamu "
           f"{fmt_rp(sub.get('price', 0))} jatuh tempo tanggal {sub.get('next_due_date')}. "
           f"Jangan lupa bayar atau cancel ya — Notifin")
    result = await send_whatsapp(user["phone"], msg)
    return {
        "status": "sent" if result.get("status") else "failed",
        "simulated": result.get("simulated", not wa_live()),
        "phone": user["phone"],
        "message": msg,
    }


def cycle_back(d: date, cycle: str) -> date:
    if cycle == "weekly":
        return d - timedelta(days=7)
    if cycle == "yearly":
        try:
            return d.replace(year=d.year - 1)
        except ValueError:
            return d.replace(year=d.year - 1, day=28)
    y, m = d.year, d.month - 1
    if m < 1:
        y, m = y - 1, 12
    day = min(d.day, calendar.monthrange(y, m)[1])
    return date(y, m, day)


@api_router.get("/groups/{gid}/history")
async def group_history(gid: str, user: dict = Depends(get_current_user)):
    g = await get_group_for_member(gid, user["user_id"])
    members = g.get("members", [])
    subs_docs = await db.group_subscriptions.find(
        {"group_id": gid, "deleted_at": None}, {"_id": 0}).sort("created_at", 1).to_list(200)

    out = []
    for s in subs_docs:
        s = await advance_group_sub(s)
        try:
            cur = date.fromisoformat(s["next_due_date"])
        except Exception:
            continue
        try:
            created = datetime.fromisoformat(s["created_at"]).date()
        except Exception:
            created = None
        cycle = s.get("billing_cycle", "monthly")
        periods = []
        p = cycle_back(cur, cycle)
        count = 0
        while count < 12 and (created is None or p >= created):
            key = p.isoformat()
            splits = compute_splits(s, members, period=key)
            periods.append({
                "period": key,
                "splits": splits,
                "paid_count": sum(1 for x in splits if x["paid"]),
                "member_count": len(splits),
            })
            p = cycle_back(p, cycle)
            count += 1
        if periods:
            out.append({"subscription": group_sub_public(s), "periods": periods})
    return {"history": out}


# ---------------------------------------------------------------------------
# Push notifications (Emergent managed relay)
# ---------------------------------------------------------------------------
@api_router.post("/register-push", status_code=201)
async def register_push(body: RegisterPushBody):
    try:
        resp = await _push_client.post("/api/v1/push/users/register", json=body.model_dump())
        if resp.status_code == 401:
            # Preview env has a placeholder push key; real key is injected at deploy.
            logger.info("register-push skipped: push key not active yet")
            return {"status": "skipped"}
        if resp.status_code >= 500:
            raise HTTPException(502, "Push provider unavailable")
        resp.raise_for_status()
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"register-push failed (non-blocking): {e}")
        return {"status": "skipped"}
    return {"status": "registered"}


async def send_push(recipients: List[str], data: dict, idempotency_key: Optional[str] = None) -> None:
    if not recipients:
        return
    if "title" not in data or "message" not in data:
        raise ValueError("data must include title and message")
    payload: dict = {"recipients": recipients[:100], "data": data}
    if idempotency_key:
        payload["$idempotency_key"] = idempotency_key
    resp = await _push_client.post("/api/v1/push/trigger", json=payload)
    if resp.status_code == 401:
        raise HTTPException(500, "EMERGENT_PUSH_KEY missing or invalid")
    if resp.status_code >= 500:
        raise HTTPException(502, "Push provider unavailable")
    resp.raise_for_status()


@api_router.get("/")
async def root():
    return {"message": "Notifin API", "status": "ok"}


# ---------------------------------------------------------------------------
# Admin — password-gated internal tool (not part of the public app) for
# manually flipping any account's plan. Useful for granting premium to a
# tester, or fixing a case where a real payment came through but Mayar's
# webhook was missed. Served at GET /admin as a plain HTML page with its
# own login; the API endpoints below sit under /api like everything else.
# ---------------------------------------------------------------------------
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")


def create_admin_token() -> str:
    payload = {"admin": True, "exp": datetime.now(timezone.utc) + timedelta(hours=12)}
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALG)


async def require_admin(authorization: Optional[str] = Header(None)) -> None:
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(status_code=401, detail="Missing admin token")
    token = authorization.split(" ", 1)[1].strip()
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALG])
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Sesi admin sudah habis, login lagi")
    if not payload.get("admin"):
        raise HTTPException(status_code=401, detail="Token bukan token admin")


class AdminLoginBody(BaseModel):
    password: str


@api_router.post("/admin/login")
async def admin_login(body: AdminLoginBody):
    if not ADMIN_PASSWORD.strip():
        raise HTTPException(
            status_code=503,
            detail="Password admin belum diatur di server (env var ADMIN_PASSWORD kosong)",
        )
    if not secrets.compare_digest(body.password, ADMIN_PASSWORD):
        raise HTTPException(status_code=401, detail="Password salah")
    return {"token": create_admin_token()}


async def enrich_users(docs: List[dict]) -> List[dict]:
    """Shape raw user docs into the admin-facing view, with each user's
    active subscription count attached. Shared by the on-screen list and
    the Excel export so both always show identical data."""
    sub_counts = await asyncio.gather(
        *[
            db.subscriptions.count_documents({"user_id": d["user_id"], "deleted_at": None})
            for d in docs
        ]
    )
    return [
        {
            "user_id": d["user_id"],
            "email": d.get("email"),
            "name": d.get("name"),
            "plan": d.get("plan", "free"),
            "phone": d.get("phone"),
            "created_at": d.get("created_at"),
            "premium_since": d.get("premium_since"),
            "premium_expires_at": d.get("premium_expires_at"),
            "last_active_at": d.get("last_active_at"),
            "subscription_count": count,
        }
        for d, count in zip(docs, sub_counts)
    ]


@api_router.get("/admin/users")
async def admin_list_users(query: str = "", _: None = Depends(require_admin)):
    q = query.strip()
    filt: dict = {}
    if q:
        safe_q = re.escape(q)
        filt = {"$or": [
            {"email": {"$regex": safe_q, "$options": "i"}},
            {"name": {"$regex": safe_q, "$options": "i"}},
        ]}
    docs = await db.users.find(filt, {"_id": 0}).sort("created_at", -1).to_list(50)
    return {"users": await enrich_users(docs)}


class AdminSetPlanBody(BaseModel):
    user_id: str
    plan: str  # "free" | "premium"


@api_router.post("/admin/set-plan")
async def admin_set_plan(body: AdminSetPlanBody, _: None = Depends(require_admin)):
    if body.plan not in ("free", "premium"):
        raise HTTPException(status_code=422, detail='plan harus "free" atau "premium"')
    update: dict = {"plan": body.plan}
    if body.plan == "premium":
        # Manual grants from this panel always run 30 days from the moment
        # you click — there's no real Mayar expiry date to use here.
        update["premium_since"] = now_utc().isoformat()
        update["premium_expires_at"] = (now_utc() + timedelta(days=DEFAULT_PREMIUM_DAYS)).isoformat()
    else:
        update["premium_expires_at"] = None
    res = await db.users.update_one({"user_id": body.user_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    logger.info(f"Admin set plan: user={body.user_id} -> {body.plan}")
    return {"status": "ok"}


def _excel_dt(value: Optional[str]) -> Optional[datetime]:
    """Parse a stored ISO timestamp into a naive datetime for Excel — Excel
    doesn't understand timezone-aware datetimes, so drop the tzinfo (values
    are stored in UTC throughout this app)."""
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(value)
        return dt.replace(tzinfo=None) if dt.tzinfo else dt
    except Exception:
        return None


def build_users_xlsx(users: List[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Akun Notifin"

    headers = [
        "Nama", "Email", "No. WhatsApp", "Status", "Tanggal Daftar",
        "Premium Sejak", "Premium Sampai", "Jumlah Langganan", "Terakhir Aktif",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    date_cols = {5, 6, 7}   # Tanggal Daftar, Premium Sejak, Premium Sampai
    datetime_cols = {9}     # Terakhir Aktif

    for u in users:
        row = [
            u.get("name") or "-",
            u.get("email") or "-",
            ("+" + u["phone"]) if u.get("phone") else "-",
            "Premium" if u.get("plan") == "premium" else "Free",
            _excel_dt(u.get("created_at")),
            _excel_dt(u.get("premium_since")),
            _excel_dt(u.get("premium_expires_at")),
            u.get("subscription_count", 0),
            _excel_dt(u.get("last_active_at")),
        ]
        ws.append(row)
        r = ws.max_row
        for col in date_cols:
            ws.cell(row=r, column=col).number_format = "dd mmm yyyy"
        for col in datetime_cols:
            ws.cell(row=r, column=col).number_format = "dd mmm yyyy hh:mm"

    for col in range(1, len(headers) + 1):
        max_len = len(headers[col - 1])
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=col).value
            max_len = max(max_len, len(str(val)) if val is not None else 0)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 42)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class AdminExportBody(BaseModel):
    user_ids: Optional[List[str]] = None  # None/empty = export every account


@api_router.post("/admin/export")
async def admin_export_users(body: AdminExportBody, _: None = Depends(require_admin)):
    if body.user_ids:
        docs = await db.users.find({"user_id": {"$in": body.user_ids}}, {"_id": 0}).to_list(len(body.user_ids))
    else:
        docs = await db.users.find({}, {"_id": 0}).sort("created_at", -1).to_list(10000)
    users = await enrich_users(docs)
    xlsx_bytes = build_users_xlsx(users)
    filename = f"notifin-akun-{date.today().isoformat()}.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


ADMIN_PAGE_HTML = """<!doctype html>
<html lang="id">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<meta name="robots" content="noindex, nofollow" />
<title>Notifin Admin</title>
<style>
  * { box-sizing: border-box; }
  body {
    margin: 0; min-height: 100vh; display: flex; align-items: center; justify-content: center;
    background: #F7FAF8; color: #182924;
    font-family: -apple-system, "Segoe UI", Roboto, sans-serif;
    padding: 24px;
  }
  .card {
    width: 100%; max-width: 480px; background: #FFFFFF; border-radius: 20px;
    padding: 28px; box-shadow: 0 16px 32px -20px rgba(11,61,46,0.28);
  }
  .wide { max-width: 760px; }
  h1 { font-size: 20px; margin: 0 0 4px; }
  p.sub { color: #6B7280; font-size: 14px; margin: 0 0 20px; }
  input {
    width: 100%; padding: 12px 14px; border-radius: 12px; border: 1.5px solid #D1D5DB;
    font-size: 15px; margin-bottom: 12px; outline: none;
  }
  input:focus { border-color: #059669; }
  button {
    cursor: pointer; border: none; border-radius: 999px; font-weight: 700; font-size: 14px;
    padding: 12px 18px;
  }
  .btn-primary { width: 100%; background: #059669; color: #fff; padding: 14px; font-size: 15px; }
  .btn-primary:disabled { opacity: 0.6; cursor: default; }
  .error { color: #EF4444; font-size: 13px; margin: -6px 0 12px; min-height: 16px; }
  .row {
    padding: 14px 0;
    border-bottom: 1px solid #F3F4F6;
  }
  .row:last-child { border-bottom: none; }
  .row-top { display: flex; align-items: center; gap: 12px; }
  .row-top .info { flex: 1; min-width: 0; }
  .row .name { font-weight: 700; font-size: 14px; }
  .row .email { color: #6B7280; font-size: 13px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
  .pill {
    font-size: 11px; font-weight: 700; padding: 3px 9px; border-radius: 999px;
    text-transform: uppercase; letter-spacing: 0.04em; white-space: nowrap;
  }
  .pill-premium { background: #D1FAE5; color: #065F46; }
  .pill-free { background: #E8F0EC; color: #233B33; }
  .btn-toggle { background: #E8F0EC; color: #182924; white-space: nowrap; }
  .btn-toggle:disabled { opacity: 0.5; cursor: default; }
  .row-meta { display: flex; flex-wrap: wrap; gap: 8px; margin-top: 10px; }
  .chip {
    font-size: 12px; color: #374151; background: #F3F4F6; padding: 5px 10px 5px 5px;
    border-radius: 8px; display: flex; align-items: center; gap: 6px; white-space: nowrap;
  }
  .chip-label {
    color: #9CA3AF; font-weight: 700; font-size: 10px; text-transform: uppercase;
    letter-spacing: 0.03em; background: #fff; padding: 3px 7px; border-radius: 6px;
  }
  .chip-ok { background: #ECFDF5; color: #047857; }
  .chip-ok .chip-label { color: #059669; }
  .chip-warn { background: #FFFBEB; color: #92400E; }
  .chip-warn .chip-label { color: #B45309; }
  .chip-danger { background: #FEF2F2; color: #991B1B; }
  .chip-danger .chip-label { color: #DC2626; }
  #app { display: none; }
  .top-row { display: flex; align-items: center; justify-content: space-between; margin-bottom: 4px; }
  .logout { background: none; color: #6B7280; font-weight: 600; padding: 4px; }
  .empty { color: #6B7280; font-size: 14px; padding: 20px 0; text-align: center; }
  .toolbar {
    display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap;
    gap: 10px; margin: 4px 0 8px;
  }
  .select-all {
    display: flex; align-items: center; gap: 6px; font-size: 13px; color: #374151;
    font-weight: 600; cursor: pointer;
  }
  .select-all input { width: auto; margin: 0; }
  .toolbar-actions { display: flex; gap: 8px; }
  .btn-export {
    background: #E8F0EC; color: #182924; font-size: 13px; padding: 9px 14px;
  }
  .btn-export:disabled { opacity: 0.5; cursor: default; }
  .btn-export-all { background: #059669; color: #fff; }
  .row-checkbox { width: auto; margin: 0; flex-shrink: 0; }
</style>
</head>
<body>

<div class="card" id="login-card">
  <h1>Notifin Admin</h1>
  <p class="sub">Masuk untuk atur status Premium akun secara manual.</p>
  <input id="password" type="password" placeholder="Password admin" onkeydown="if(event.key==='Enter')login()" />
  <div class="error" id="login-error"></div>
  <button class="btn-primary" id="login-btn" onclick="login()">Masuk</button>
</div>

<div class="card wide" id="app">
  <div class="top-row">
    <h1>Notifin Admin</h1>
    <button class="logout" onclick="logout()">Keluar</button>
  </div>
  <p class="sub">Cari akun berdasarkan email atau nama, lalu ubah status Premium-nya.</p>
  <input id="search" type="text" placeholder="Cari email atau nama..." oninput="onSearchInput()" />
  <div class="error" id="app-error"></div>
  <div class="toolbar">
    <label class="select-all">
      <input type="checkbox" id="select-all-checkbox" onchange="onSelectAll(this.checked)" />
      Pilih semua
    </label>
    <div class="toolbar-actions">
      <button class="btn-export" id="export-selected-btn" onclick="exportSelected()" disabled>
        Export Terpilih (0)
      </button>
      <button class="btn-export btn-export-all" onclick="exportAll()">Export Semua (.xlsx)</button>
    </div>
  </div>
  <div id="list"></div>
</div>

<script>
  let token = null;
  let searchTimer = null;
  let selected = new Set();
  let currentUsers = [];

  async function login() {
    const password = document.getElementById('password').value;
    const errEl = document.getElementById('login-error');
    const btn = document.getElementById('login-btn');
    errEl.textContent = '';
    btn.disabled = true;
    try {
      const res = await fetch('/api/admin/login', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ password }),
      });
      const data = await res.json();
      if (!res.ok) { errEl.textContent = data.detail || 'Gagal masuk'; return; }
      token = data.token;
      document.getElementById('login-card').style.display = 'none';
      document.getElementById('app').style.display = 'block';
      loadUsers('');
    } catch (e) {
      errEl.textContent = 'Tidak bisa menghubungi server.';
    } finally {
      btn.disabled = false;
    }
  }

  function logout() {
    token = null;
    document.getElementById('app').style.display = 'none';
    document.getElementById('login-card').style.display = 'block';
    document.getElementById('password').value = '';
  }

  function onSearchInput() {
    clearTimeout(searchTimer);
    const q = document.getElementById('search').value;
    searchTimer = setTimeout(() => loadUsers(q), 300);
  }

  async function loadUsers(query) {
    const errEl = document.getElementById('app-error');
    errEl.textContent = '';
    try {
      const res = await fetch('/api/admin/users?query=' + encodeURIComponent(query), {
        headers: { Authorization: 'Bearer ' + token },
      });
      const data = await res.json();
      if (!res.ok) {
        if (res.status === 401) { logout(); return; }
        errEl.textContent = data.detail || 'Gagal memuat daftar akun';
        return;
      }
      renderList(data.users);
    } catch (e) {
      errEl.textContent = 'Tidak bisa menghubungi server.';
    }
  }

  function fmtDate(iso) {
    if (!iso) return '-';
    const d = new Date(iso);
    if (isNaN(d.getTime())) return '-';
    return d.toLocaleDateString('id-ID', { day: 'numeric', month: 'short', year: 'numeric' });
  }

  function renewalInfo(expiresAt) {
    if (!expiresAt) return { text: '-', cls: '' };
    const d = new Date(expiresAt);
    if (isNaN(d.getTime())) return { text: '-', cls: '' };
    const days = Math.ceil((d.getTime() - Date.now()) / 86400000);
    if (days < 0) return { text: 'Lewat ' + Math.abs(days) + ' hari', cls: 'chip-danger' };
    if (days === 0) return { text: 'Hari ini', cls: 'chip-danger' };
    if (days <= 7) return { text: days + ' hari lagi', cls: 'chip-warn' };
    return { text: days + ' hari lagi', cls: 'chip-ok' };
  }

  function lastActiveText(iso) {
    if (!iso) return 'Belum pernah';
    const d = new Date(iso);
    if (isNaN(d.getTime())) return 'Belum pernah';
    const mins = Math.round((Date.now() - d.getTime()) / 60000);
    if (mins < 1) return 'Baru saja';
    if (mins < 60) return mins + ' menit lalu';
    const hours = Math.round(mins / 60);
    if (hours < 24) return hours + ' jam lalu';
    return Math.round(hours / 24) + ' hari lalu';
  }

  function chip(label, value, cls) {
    return '<div class="chip ' + (cls || '') + '"><span class="chip-label">' + label + '</span>' + escapeHtml(String(value)) + '</div>';
  }

  function toggleSelect(uid, checked) {
    if (checked) selected.add(uid); else selected.delete(uid);
    updateExportUi();
  }

  function onSelectAll(checked) {
    currentUsers.forEach((u) => {
      if (checked) selected.add(u.user_id); else selected.delete(u.user_id);
    });
    renderList(currentUsers);
  }

  function updateExportUi() {
    const btn = document.getElementById('export-selected-btn');
    btn.textContent = 'Export Terpilih (' + selected.size + ')';
    btn.disabled = selected.size === 0;
    const selectAllCb = document.getElementById('select-all-checkbox');
    if (selectAllCb) {
      selectAllCb.checked = currentUsers.length > 0 && currentUsers.every((u) => selected.has(u.user_id));
    }
  }

  async function downloadXlsx(userIds) {
    const errEl = document.getElementById('app-error');
    errEl.textContent = '';
    try {
      const res = await fetch('/api/admin/export', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json', Authorization: 'Bearer ' + token },
        body: JSON.stringify({ user_ids: userIds }),
      });
      if (!res.ok) {
        if (res.status === 401) { logout(); return; }
        const data = await res.json().catch(() => ({}));
        errEl.textContent = data.detail || 'Gagal export data';
        return;
      }
      const blob = await res.blob();
      const url = URL.createObjectURL(blob);
      const a = document.createElement('a');
      a.href = url;
      a.download = 'notifin-akun-' + new Date().toISOString().slice(0, 10) + '.xlsx';
      document.body.appendChild(a);
      a.click();
      a.remove();
      setTimeout(() => URL.revokeObjectURL(url), 1000);
    } catch (e) {
      errEl.textContent = 'Tidak bisa menghubungi server.';
    }
  }

  function exportAll() {
    downloadXlsx(null);
  }

  function exportSelected() {
    if (selected.size === 0) return;
    downloadXlsx(Array.from(selected));
  }

  function renderList(users) {
    currentUsers = users;
    const list = document.getElementById('list');
    if (!users.length) {
      list.innerHTML = '<div class="empty">Tidak ada akun ditemukan.</div>';
      updateExportUi();
      return;
    }
    list.innerHTML = users.map((u) => {
      const isPremium = u.plan === 'premium';
      const renewal = isPremium ? renewalInfo(u.premium_expires_at) : null;
      let meta = chip('Daftar', fmtDate(u.created_at));
      if (isPremium) {
        meta += chip('Premium sejak', fmtDate(u.premium_since));
        meta += chip('Renew', renewal.text, renewal.cls);
      }
      meta += chip('WA', u.phone ? '+' + u.phone : '-');
      meta += chip('Langganan', u.subscription_count);
      meta += chip('Aktif', lastActiveText(u.last_active_at));
      const checked = selected.has(u.user_id) ? 'checked' : '';

      return (
        '<div class="row">' +
          '<div class="row-top">' +
            '<input type="checkbox" class="row-checkbox" data-uid="' + u.user_id + '" ' + checked + ' onchange="toggleSelect(this.getAttribute(&quot;data-uid&quot;), this.checked)" />' +
            '<div class="info">' +
              '<div class="name">' + escapeHtml(u.name || '(tanpa nama)') + '</div>' +
              '<div class="email">' + escapeHtml(u.email || '') + '</div>' +
            '</div>' +
            '<span class="pill ' + (isPremium ? 'pill-premium' : 'pill-free') + '">' +
              (isPremium ? 'Premium' : 'Free') +
            '</span>' +
            '<button class="btn-toggle" data-uid="' + u.user_id + '" data-plan="' + (isPremium ? 'free' : 'premium') + '" onclick="togglePlan(this)">' +
              (isPremium ? 'Jadikan Free' : 'Jadikan Premium') +
            '</button>' +
          '</div>' +
          '<div class="row-meta">' + meta + '</div>' +
        '</div>'
      );
    }).join('');
    updateExportUi();
  }

  async function togglePlan(btn) {
    const userId = btn.getAttribute('data-uid');
    const plan = btn.getAttribute('data-plan');
    btn.disabled = true;
    const original = btn.textContent;
    btn.textContent = 'Menyimpan...';
    try {
      const res = await fetch('/api/admin/set-plan', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json', Authorization: 'Bearer ' + token },
        body: JSON.stringify({ user_id: userId, plan }),
      });
      const data = await res.json();
      if (!res.ok) {
        document.getElementById('app-error').textContent = data.detail || 'Gagal menyimpan';
        btn.disabled = false;
        btn.textContent = original;
        return;
      }
      loadUsers(document.getElementById('search').value);
    } catch (e) {
      document.getElementById('app-error').textContent = 'Tidak bisa menghubungi server.';
      btn.disabled = false;
      btn.textContent = original;
    }
  }

  function escapeHtml(s) {
    const div = document.createElement('div');
    div.textContent = s;
    return div.innerHTML;
  }
</script>
</body>
</html>
"""


@app.get("/admin", response_class=HTMLResponse)
async def admin_page():
    return HTMLResponse(ADMIN_PAGE_HTML)


# ---------------------------------------------------------------------------
# Startup: indexes
# ---------------------------------------------------------------------------
@app.on_event("startup")
async def startup():
    try:
        await db.users.create_index("email", unique=True)
        await db.users.create_index("user_id", unique=True)
        await db.user_sessions.create_index("session_token", unique=True)
        await db.user_sessions.create_index("user_id")
        await db.subscriptions.create_index("user_id")
        await db.subscriptions.create_index("id", unique=True)
        await db.groups.create_index("id", unique=True)
        await db.groups.create_index("invite_code", unique=True)
        await db.groups.create_index("members.user_id")
        await db.group_subscriptions.create_index("id", unique=True)
        await db.group_subscriptions.create_index("group_id")
        await db.notif_log.create_index("key", unique=True)
        await db.spending_snapshots.create_index(
            [("user_id", 1), ("period", 1)], unique=True)
        await db.mayar_webhook_log.create_index("received_at")
    except Exception as e:
        logger.warning(f"index creation: {e}")
    asyncio.create_task(scheduler_loop())


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[
        "https://emergent-app-vert.vercel.app",
        "https://notifin.online",
        "https://www.notifin.online",
    ],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
    await _push_client.aclose()
